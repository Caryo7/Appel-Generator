import openpyxl as xl

# A ce stade, on a de quoi extraire le colloscope du fichier excel.
# On peut récupérer la liste des colles de chaque groupes, et leur POSIX
# On peut en faire de même pour chaque prof.
# Il reste alors à faire un programme qui lit l'emploit du temps des élèves
# Et qui insert aux bons endroits les cours.
# Pour faire plus simple, on supposera que les horraires de tout le monde
# Sont les mêmes, et que seules les matières changent.
# Alors, on peut écrire

class EDT:
    def __init__(self, wb):
        self.wb = wb
        self.sh = self.wb.active
        
    def me(self, groupe):
        pass
    
    def feed(self, groupe_id):
        pass
    
    def fill(self, colle):
        pass

    def export(self, folder):
        for col in range(1, 20): # range du nombre de colonnes de l'EDT
            for row in range(1, 40): # range du nombre de lignes de l'EDT
                v = self.sh.cell(column=col, row = row).value
                
                if v == '$id':
                    self.sh.cell(column=col, row = row).value = None


def import_edt(path):
    """Ouvre l'emploi du temps des élèves, et en fait une "copie"
    Arguments
    * path : le lien vers le fichier qui contient l'emploi du temps
             au format excel.
             
    Retourne un emploi du temps (class EDT)
    """
    
    wb = xl.load_workbook(path)
    e = EDT(wb)
    
    return e

def fill_edt(groupes, edt, folder):
    """Avec un dictionnaire des groupes de colle et un emploi du temps,
    vient remplir les trous volontairement laissés par le professeur
    en charge de la création des emplois du temps.
    Chaque groupe de colle à donc un emploi du temps qui lui est propre.
    Sous réserve bien sur des hypothèses simplificatrices susmentionnées
    plus haut à propos des cours qui ont lieux en même temps.
    
    Arguments
    *  groupes : dictionnaires (groupe: colles)
    * edt      : emploi du temps (class EDT)
    * folder   : dossier de sortie des emplois du temps générés
    """
    
    for groupe, semaine in groupes.items():
        groupe_id = semaine.groupe_id
        edt.me(groupe)
        edt.feed(groupe_id)
        
        for colle in semaine.colles:
            edt.fill(colle)
            
        edt.export(folder)
