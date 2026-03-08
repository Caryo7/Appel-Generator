import openpyxl as xl
from openpyxl.styles import PatternFill
import os
from zipfile import *
from pathlib import Path
from datetime import datetime

import excelsaver
import dialogs
import config

import box

IDLE_MODE = config.idleMode() # or '\n' if console
WHITE_FILL = PatternFill()

# A ce stade, on a de quoi extraire le colloscope du fichier excel.
# On peut récupérer la liste des colles de chaque groupes, et leur POSIX
# On peut en faire de même pour chaque prof.
# Il reste alors à faire un programme qui lit l'emploit du temps des élèves
# Et qui insert aux bons endroits les cours.
# Pour faire plus simple, on supposera que les horraires de tout le monde
# Sont les mêmes, et que seules les matières changent.
# Alors, on peut écrire

def dt(h1, h2):
    h1, m1 = h1.split('h')
    h2, m2 = h2.split('h')
    h1 = int(h1)
    h2 = int(h2)
    m1 = int(m1) if m1 else 0
    m2 = int(m2) if m2 else 0

    dh = abs(h2-h1)
    dm = abs(m2-m1)
    return dh == 0 and dm <= 15


class EDT:
    def __init__(self, path, name, lang, title):
        self.wb = xl.load_workbook(path)
        self.colles = []
        self.name = name
        self.lang = lang
        self.title = title

    def me(self, groupe, semaine, colle):
        self.groupe = groupe
        for col in range(1, 40): # range du nombre de colonnes de l'EDT
            for row in range(1, 30): # range du nombre de lignes de l'EDT
                v = self.sh.cell(column=col, row = row).value
                if v is None:
                    continue

                if v.upper() == 'CLASSE':
                    self.sh.cell(column=col, row=row).value = self.title.format(groupe = groupe,
                                                                                semaine = semaine,
                                                                                colle = colle,)
                elif v.upper() == 'DS':
                    self.sh.cell(column = col, row = row).value = f'DS\n{semaine.DS}'

                if 'LV2' in v.upper():
                    if self.lang is not None:
                        self.sh.cell(column = col, row = row).value = self.lang.capitalize()

                    else:
                        self.sh.cell(column = col, row = row).value = ''
                        self.sh.cell(column = col, row = row).fill = WHITE_FILL

    def feed(self, groupe_id, ssgrp_id):
        groupe_id = groupe_id.replace('/', '-')
        ssgrp = str('-' + ssgrp_id) if ssgrp_id is not None else ''
        n = groupe_id + ssgrp
        self.sh = self.wb[n]
        for sh in list(self.wb):
            if sh.title == n:
                continue

            del self.wb[sh.title]

    def fill(self, colle):
        day_line = 3
        col_heure = 1
        for col in range(1, 40): # range du nombre de colonnes de l'EDT
            for row in range(1, 30): # range du nombre de lignes de l'EDT
                v = self.sh.cell(column=col, row = row).value
                if v is None:
                    continue

                if v.upper() == 'COLLE':
                    heure_edt = self.sh.cell(column=col_heure, row = row).value.lower().split('-')[0]
                    jour_edt = self.sh.cell(column = col, row = day_line).value.lower()
                    heure_cl = colle.heure.lower().replace('00', '').split('-')[0]
                    jour_cl = colle.jour.lower()

                    if jour_edt == jour_cl and dt(heure_cl, heure_edt):
                        self.sh.cell(column=col, row = row).value = '{colleur} {salle}'.format(
                            salle = colle.salle,
                            colleur = colle.prof)

                        return True

        return False

    def export(self, folder):
        for col in range(1, 40): # range du nombre de colonnes de l'EDT
            for row in range(1, 30): # range du nombre de lignes de l'EDT
                v = self.sh.cell(column=col, row = row).value
                if v is None:
                    continue

                if v.upper() == 'COLLE':
                    self.sh.cell(column=col, row = row).value = None

                elif v.upper() == 'DATE':
                    self.sh.cell(column=col, row = row).value = datetime.today().strftime("%d/%m/%y")

                elif v.upper() == 'HEURE':
                    self.sh.cell(column=col, row = row).value = datetime.today().strftime("%H:%M")

        fp = os.path.join(os.path.abspath('.'), folder, 'groupe-{}-{}'.format(self.groupe, self.name))
        self.wb.save(fp + '.xlsx')
        return excelsaver.export_pdf(fp + '.xlsx')


def zip_output(paths, semaine, folder):
    """Une fois l'enregistrement terminé, on peut mettre dans un fichier ZIP
    l'intégralité des EDT de sortie. Pour cela,
    Arguments
    * paths  : la liste des fichiers à mettre en ZIP
    * semaine: le numéro de la semaine

    Ne retourne rien
    """

    z = ZipFile(f'./{folder}/output-S{semaine}.zip', 'w', compression=ZIP_DEFLATED, compresslevel=9)
    for fp in paths:
        f = open(fp, 'rb')
        data = f.read()
        f.close()

        fz = z.open(fp.split('/')[-1], 'w')
        fz.write(data)
        fz.close()

    z.close()

def fill_edt(groupes, path, folder, semaine_nb, table_addr, config):
    """Avec un dictionnaire des groupes de colle et un emploi du temps,
    vient remplir les trous volontairement laissés par le professeur
    en charge de la création des emplois du temps.
    Chaque groupe de colle à donc un emploi du temps qui lui est propre.
    Sous réserve bien sur des hypothèses simplificatrices susmentionnées
    plus haut à propos des cours qui ont lieux en même temps.
    
    Arguments
    * groupes : dictionnaires (groupe: colles)
    * path    : Fichier emploi du temps
    * folder  : dossier de sortie des emplois du temps générés
    * semaine : le numéro de la semaine en cours
    * table_addr: la table des élèves avec leurs adresses mails, ...
    * config  : la configuration 
    """

    pc = 0.0
    opc = 0.0
    fps = []
    ce = []
    lm = 0
    for groupe, semaine in groupes.items():
        for nom, family, _, lang, ssgrp in table_addr[groupe]:
            if len(nom) > lm:
                lm = len(nom)

    p = box.Progress('Compilation...', length=3*len(groupes), larg = lm)
    for groupe, semaine in groupes.items():
        groupe_id = semaine.groupe_id
        for nom, family, _, lang, ssgrp in table_addr[groupe]:
            edt = EDT(path, nom+'.'+family, lang, config.edt_title)
            edt.feed(groupe_id, ssgrp)
            edt.me(groupe, semaine.colles[0].semaine, semaine.colles[0].colle_id) # On récupère le numéro de la semaine via la première colle du groupe

            for colle in semaine.colles:
                r = edt.fill(colle)
                if not r:
                    ce.append(f'{str(colle)}, {colle.jour}, {colle.heure}, {colle.prof}, {colle.salle}')

            r = edt.export(folder)
            if not r:
                return False

            fps.append(r)
            p.step(nom)

    if ce:
        box.warning('Colision de colles !', ce)

    zip_output(fps, semaine_nb, folder)
    return True

def clear():
    for fp in list(Path('output/').glob('**/*.xlsx')):
        os.remove(fp)
