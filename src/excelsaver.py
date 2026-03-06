from win32com.client import DispatchEx
from openpyxl.styles import PatternFill
from openpyxl.styles.fonts import Font
import openpyxl as xl
import dialogs
import config as confr # On utilise le fichier config du dossier

excel = DispatchEx('Excel.Application')
excel.Visible = 0

def export_pdf(path_from, path_to = None):
    """Enregistrement du fichier excel. Utilise une
    bibliothèque de control de Excel pour demander à excel de
    faire le travail.

    Arguments
    * path_from: le fichier excel à exporter
    * path_to: le fichier PDF de sortie

    Retourne
    * path_to
    """

    if path_to is None:
        # Si on ne donne pas de destination, on prendra le même
        # nom que le fichier excel
        path_to = path_from.replace('.xlsx', '.pdf')

    try:
        wb = excel.Workbooks.Open(path_from)
        wb.application.displayalerts = False
        ws = wb.Worksheets[0]
        ws.SaveAs(path_to, FileFormat=57)
        wb.Close()

    except Exception as e:
        # Si erreur d'exportation on retournera faux, pour erreur
        dialogs.warning("Erreur d'exportation ! Fermez les taches Excel déjà démarrées")
        excel.Quit()
        return False

    return path_to


def appel(tables, path, config, week = None):
    """Création d'une feuille d'appel pour chaque groupes pour chaque
    semaine. Cela utilise une table sur une semaine !

    Arguments
    * tables : un dictionnaire de table, une table par semaine
    * path  : fichier excel à générer
    * config : le fichier de configuration des feuilles
    * week : optionnel pour indiquer une semaine en particulier.
             Si semaine est donné, il n'y aura que la feuille en question !

    Retourne
    Rien
    """

    feuilles = config.feuilles
    wb = xl.Workbook()
    wb.remove(wb.active)
    font_student = Font()
    font_title = Font(bold = True)
    fill_student = PatternFill()#start_color='bad9ff', end_color='bad9ff', fill_type="solid")
    fill_title = PatternFill()#start_color='bad9ff', end_color='bad9ff', fill_type="solid")
    sheet_layout = config.layout

    for semaine, table in tables.items():
        if week != semaine and week is not None:
            continue

        sh = wb.create_sheet('Semaine-{}'.format(semaine))
        sh.cell(row = 1, column = 1).value = 'Appel semaine ' + str(semaine)
        sh.cell(row = 1, column = 1).font = font_title
        sh.cell(row = 1, column = 1).fill = fill_title
        sh.page_margins.left = 1
        sh.page_margins.right = 1
        sh.page_margins.top = 1
        sh.page_margins.bottom = 1
        sh.sheet_properties.pageSetUpPr.fitToPage = True

        data = {i: [] for i in range(len(feuilles))}
        for colle in table:
            for i, sheet in enumerate(feuilles):
                if colle.colle_id in sheet[0]:
                    data[i].append(colle)

        extract = {}
        eleves = []
        for i, colles in data.items():
            extract[i] = []
            for colle in colles:
                if colle.eleves is None:
                    continue

                for eleve in colle.eleves:
                    if eleve not in extract[i]:
                        extract[i].append(eleve)
                        if len(eleve) not in eleves:
                            eleves.append(len(eleve))

        for k in extract:
            extract[k].sort() # Tri alphabétique des noms dans la feuille d'appel

        for i, (_, title) in enumerate(feuilles):
            col, row_base = sheet_layout[i]
            sh.cell(row = row_base, column = col).value = title
            sh.cell(row = row_base, column = col).font = font_title
            sh.cell(row = row_base, column = col).fill = fill_title

            for line, student in enumerate(extract[i]):
                sh.cell(row = line + row_base + 1, column = col).value = student
                sh.cell(row = line + row_base + 1, column = col).font = font_student
                sh.cell(row = line + row_base + 1, column = col).fill = fill_student

            sh.column_dimensions[chr(col+64)].width = max(eleves + [len(title)]) * 1.2

    wb.save(path)

    if week is not None or 1:
        export_pdf(os.path.abspath(path))
