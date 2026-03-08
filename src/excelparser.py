## Générateur de liste d'appel
## Programme écrit par Benoit Charreyron
## Copyright 2025
## MIT License
## Fichier principal de traitement interne

from pathlib import Path
import sys
import os
import openpyxl as xl

import config as confr # On utilise le fichier config du dossier
import dialogs
import edtfiller
import excelsaver

import box

class Colle:
    # Attributs de la classe (fixe pour chaque colle)
    salle = None
    heure = None
    jour = None
    semaine = None
    
    # Attributs variables (dépend des colles)
    prof = None # Champs remplis plus tard !
    eleves = None
    groupe = None
    
    # Champ spéficique pour l'élève
    colle_id = None
    
    def __init__(self,
                 salle,
                 heure,
                 jour,
                 semaine,
                 prof,
                 groupe,
                 colle_id,
                 ):
        """Initialisation de la classe colle.
        Arguments
        * salle : la salle de la colle
        * heure : l'heure de la colle
        * jour : le jour dans la semaine de la colle
        * semaine : le numéro de semaine de la colle
        * prof : le professeur qui fait la colle
        * groupe : le groupe qui suit la colle
        * colle_id : l'identifiant de la colle (et du groupe cette semaine là)

        Retourne
        * Colle : la colle
        """

        self.salle = salle
        self.heure = heure
        self.jour = jour
        self.semaine = semaine
        self.prof = prof
        self.groupe = groupe
        self.colle_id = colle_id

    def __repr__(self):
        return '{groupe}-{colle_id}'.format(
            groupe = self.groupe,
            colle_id = self.colle_id
            )

def create_groups(table):
    """Extrait de la table des adresses mail la liste des groupes
    et leurs élèves.

    Arguments
    * table : la table des adresse mail venant de automail.py

    Retourne
    * groupes : dict des groupes/noms
    """

    grp = {}
    for groupe, k in table.items():
        ns = []
        for nom, family, addr, lang, ssgrp in k:
            ns.append(nom)

        grp[groupe] = ns

    return grp

def read_colloscope(path, config, table_addr):
    """Lecture du fichier excel .xlsx du colloscope.
    Ce dernier sera traité selon la configuration, et retournera une
    liste de colles. Cette liste de colle sera mélangée, mais chaque colle
    aura sont propre identifiant pour se repérer dans le temps !

    Arguments
    * path : fichier excel à ouvrir
    * config : configuration actuelle
    * table_addr : la table des emails pour avoir tous les groupes !
    
    Sortie
    * table : liste de Colle
    """

    # Section ouverture du fichier excel
    wb = xl.load_workbook(path)
    sh = box.ask_feuille("Ouverture du colloscope", wb, path, config.default_col_sheet)

    # Section lecture de la liste des élèves
    groupes =  create_groups(table_addr)

    # Section lecture du colloscope

    # Paramétrage interne de la fonction. Ce dernier sera a terme
    # sauvegardé dans un fichier de configuration .ini.
    # Voir pour avoir les droits d'écriture dans le dossier de travail.
    col_prof = config.col_prof # Format int
    col_salle = config.col_salle # "
    col_heure = config.col_heure # "
    col_jour = config.col_jour # "
    col_id = config.col_id # " Colle identifiant pour savoir dans quel groupe on
                # est pour le reste de l'emploi du temps !

    col_groupes = config.col_groupes # Format list
    
    lignes_semaine = config.lignes_semaine # Format int
    lignes = config.lignes # Format list

    # Le choix est fait de parcourir toutes les lignes en premier,
    # On aurait aussi pu parcourir les colonnes en premier
    table = []
    ws = []
    for line in lignes:
        prof = sh.cell(column=col_prof, row=line).value
        salle = sh.cell(column=col_salle, row=line).value
        heure = sh.cell(column=col_heure, row=line).value
        jour = sh.cell(column=col_jour, row=line).value

        l = line
        while sh.cell(column = col_id, row=l).value is None:
            l -= 1
        cid = sh.cell(column=col_id, row=l).value

        #print(prof, salle, heure, jour, cid)

        for line_semaine in lignes_semaine:
            for col in col_groupes:
                semaine = sh.cell(column=col, row = line_semaine).value
                if semaine is None:
                    continue

                l = line
                while sh.cell(column = col, row=l).value is None:
                    l -= 1

                groupe = sh.cell(column = col, row=l).value
                #print('   ', semaine, groupe)
                Kh = Colle(
                    salle,
                    heure,
                    jour,
                    semaine,
                    prof,
                    groupe,
                    cid)

                try:
                    Kh.eleves = groupes[groupe]

                except:
                    ws.append(f'Attention {prof} semaine {semaine} a le groupe {groupe} qui est inconnu !')

                table.append(Kh)

    wb.close() # Ne pas oublier pour ne pas bloquer excel !
    if ws:
        box.warning("Reconnaissance des colles", ws)

    return table

def get_this_ds(path, config, semaine):
    """Récupération du DS de la semaine dans le colloscope.
    Ouvre le colloscope et vient lire la table spéciale des DS
    Arguments
    * path: le fichier excel du colloscope
    * config: le fichier de configuration
    * semaine: la semaine actuelle

    Retourne
    * un nom de DS
    """

    wb = xl.load_workbook(path)
    sh = box.ask_feuille('Détection du DS', wb, path, default = config.ds_sheet)
    col = config.col_sem_ds
    col_ds = config.col_ds
    lignes = config.lignes_ds
    for l in lignes:
        if sh.cell(column = col, row = l).value == semaine:
            return sh.cell(column = col_ds, row = l).value

def all_weeks(table):
    """Cette fonction prend en argument une table et sort toutes les semaines
    qui y sont présentés"""

    weeks = []
    for colle in table:
        if colle.semaine not in weeks:
            weeks.append(colle.semaine)

    return weeks

def selector(table, semaine):
    """Cette fonction sera à extraire de la liste complète des colles
    uniquement les colles de la semaine intéresée.
    Le principe est le parcours général de la liste, pour l'extraction.
    
    Arguments
    * table : liste de colle de la fonction read_colloscope
    * semaine : identifiant de la semaine (ATTENTION, en str)
     
    Retourne
    * table : liste de colle de la semaine !
    """
    
    extraction = [] # On pourrait le faire par compréhension, mais c'est 
                    # plus indigeste à écrire
    
    for colle in table:
        if colle.semaine == semaine:
            extraction.append(colle)
            extraction[-1].semaine = semaine
            
    return extraction

def find_colle_id(table, groupe):
    """Cette fonction retourne l'identifiant
    du groupe de colle en à partir d'un nom de groupe
    et d'une table de groupes pour une semaine

    Arguments
    * table : la table des colles de la semaine
    * groupe : le groupe dont on cherche l'identifiant

    Retourne
    * colle_id : identifiant de colle si trouvée, None sinon
    """

    for colle in table:
        if colle.groupe.lower() == groupe.lower():
            return colle.colle_id

def read_modifs(modif_file, table):
    """Cette fonction sert à modifier le colloscope principal
    en changeant des colles et/ou en ajoutant / supprimant.
    Cela permet de manipuler la dernière semaine

    Arguments
    * modif_file : le chemin du fichier contenant les modifications
                   (qui doit être sur une seule semaine (après selector)
    * table : la table principale des groupes

    Retourne
    * table : table principal qui a été modifiée
    """

    wb = xl.load_workbook(modif_file)
    sh = box.ask_feuille("Modifications de dernière minute", wb, modif_file)

    ## Formatage
    # Colonne 1 : Groupe
    # Colonne 2 : Jour
    # Colonne 3 : Heure
    # Colonne 4 : Professeur
    # Colonne 5 : Salle
    # Première ligne : 2

    modifs = []
    row = 1
    colle_id = None
    try:
        semaine = table[0].semaine
    except:
        return None

    while sh.cell(row = row+1, column = 1).value:
        row += 1
        groupe = sh.cell(row = row, column = 1).value
        jour = sh.cell(row = row, column = 2).value
        heure = sh.cell(row = row, column = 3).value
        prof = sh.cell(row = row, column = 4).value
        room = sh.cell(row = row, column = 5).value

        kolle = Colle(room, heure.lower(), jour, semaine, prof, groupe, colle_id)
        modifs.append(kolle)

    output_table = []
    used = []
    for colle_org in table:
        updated = False
        for colle_mod in modifs:
            if colle_mod in used:
                continue

            if colle_org.jour.lower() != colle_mod.jour.lower():
                continue

            if not edtfiller.dt(colle_org.heure.lower(), colle_mod.heure.lower()):
                continue

            if colle_org.groupe.lower() != colle_mod.groupe.lower():
                continue

            colle_org.salle = colle_mod.salle
            colle_org.prof = colle_mod.prof
            output_table.append(colle_org)
            used.append(colle_mod)
            updated = True

        if not updated:
            output_table.append(colle_org)


    for colle_mod in modifs:
        if colle_mod in used:
            continue

        colle_id = find_colle_id(table, colle_mod.groupe)
        colle_mod.colle_id = colle_id
        output_table.append(colle_mod)

    return output_table


class Week:
    def __init__(self, groupe_id):
        self.groupe_id = groupe_id # = colle_id
        self.colles = []
        
    def append(self, colle):
        self.colles.append(colle)

    def __repr__(self):
        return '{groupe_id}-{nbcolle}'.format(
            groupe_id = self.groupe_id,
            nbcolle = len(self.colles))


def sort_groupes(table):
    """Cette fonction prend une table sur une semaine uniquement !
    (Voir pour passer la table comùplète par selector(table, semaine)
     pour l'extraire) et retourne un dictionnaire par groupe de colle.
     Ce dernier dictionnaire contiendra la colle de la semaine.
     Si le groupe possède plusieurs colles, il faut voir si on met une 
     liste...
     
     Arguments
     * table : une liste de colle pour UNE semaine
     
     Retourne
     * dictionnaire: (groupe: colles (class Week))
     """

    groupes = {}
    for colle in table:
        if colle.groupe not in groupes:
            groupes[colle.groupe] = Week(colle.colle_id)

        groupes[colle.groupe].append(colle)

    return groupes

def sort_profs(table):
    """Cette fonction prend une table sur une semaine uniquement !
    (Voir pour passer la table comùplète par selector(table, semaine)
     pour l'extraire) et retourne un dictionnaire par colleur.
     Ce dernier dictionnaire contiendra la colle de la semaine.
     Si le colleurfait plusieurs colles, il faut voir si on met une 
     liste...
     
     Arguments
     * table : une liste de colle pour UNE semaine

     Retourne
     * dictionnaire: (prof: colles)
     """

    profs = {}
    for colle in table:
        if colle.prof in profs:
            profs[colle.prof].append(colle)
        else:
            profs[colle.prof] = [colle]
            
    return profs
