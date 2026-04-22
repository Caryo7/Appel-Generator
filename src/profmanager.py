import config
import box
import excelparser
import openpyxl as xl
import os
import automail

def listToStr(lst):
    if len(lst) == 1:
        return lst[0]

    txt = ''
    for l in lst[:-1]:
        txt += l + ', '

    txt = txt[:-2]
    txt += ' et ' + lst[-1]
    return txt

class ProfManager:
    sujet = 'Colles de la semaine'
    remap_table = {'A': '<span style="color: #ff0000; font-weight: bold">Fantôme !</span>'}

    def __init__(self, ems, addr, template, semaine):
        self.ems = ems
        self.semaine = semaine
        self.template = template
        self.table = {}
        self.mails = {}

        wb = xl.load_workbook(addr)
        sh = box.ask_feuille("Ouverture des adresses des professeurs", wb, addr, 0)
        row = 2
        self.lm = 0
        while sh.cell(row = row, column = 1).value:
            civilite = sh.cell(row = row, column = 1).value
            prof = sh.cell(row = row, column = 2).value
            addr = sh.cell(row = row, column = 3).value

            if len(prof) > self.lm:
                self.lm = len(prof)

            self.mails[prof] = (civilite, addr)
            row += 1

    def feed(self, thistable):
        for k, v in thistable.items():
            if k not in self.table:
                self.table[k] = excelparser.Week(v.ide)

            for colle in v.colles:
                self.table[k].append(colle)

    def start(self):
        os.system(f'edit "{self.template}"')

        f = open(self.template, 'r', encoding = 'utf-8')
        data = f.read()
        f.close()

        p = box.Progress('Envoi automatique des mails professeurs', len(self.table), larg = self.lm)
        for prof, colles in self.table.items():
            try:
                civilite, addr = self.mails[prof]
            except:
                box.warning("Mails automatiques professeurs", [f"Le professeur {prof} est inconnu dans la base"])
                addr = None

            p.step(prof, color = 'red' if not addr else None, bar = 'yellow' if not addr else None)
            if not addr:
                continue

            liste = "<ul>"
            for colle in colles.colles:
                grp = colle.groupe
                eleves = colle.eleves
                if grp in self.remap_table:
                    grp = self.remap_table[grp]
                    eleves = []

                txt = f'{colle.jour} à {colle.heure} en {colle.salle}, le groupe {grp}'
                if eleves:
                    txt += f' ({listToStr(eleves)})'

                liste += '  <li>' + txt + '</li>'

            liste += '</ul>'

            var = {'liste': liste,
                   'civilite': civilite,
                   'semaine': self.semaine,
                   }

            r = self.ems.send(addr, self.sujet,
                              automail.autoformat(data, var),
                              files = [],
                              test = False)

            if not r:
                return
