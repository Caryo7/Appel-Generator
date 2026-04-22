from tkinter.messagebox import *
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import email, smtplib, ssl
import openpyxl as xl
from pathlib import Path
from configparser import ConfigParser
import dialogs
import time
import config
import box

TEST_MODE = config.noMail()
IDLE_MODE = config.idleMode()
MAX_MAIL_NB = 20

HEADER = """<html>
  <head>
    <style type="text/css">
      p {
        color: #000000
        line-height: 1;
      }
      .warning {
        color: #ff0000;
        font-weight: bold;
      }
    </style>
  </head>
  <body>
"""

FOOTER = """  </body>
</html>"""

def plainToHtml(plain):
    plain = plain.replace('</ul>\n', '</ul>')
    plain = plain.replace('</ul>\n', '</ul>')
    html = '<p>' + plain.replace('\n', '<br>\n') + '</p>'

    return HEADER + html + FOOTER

class EmailSender:
    counter = 0

    def connect(self, mail = None, pwd = None):
        """Connexion au serveur mail au démarrage du programme.
        Plus besoin de se reconnecter à chaque mail et ainsi générer une erreur
        Arguments
        * mail : l'adresse email
        * pwd : le mot de passe de connexion

        Retourne
        Rien
        """

        if mail is not None:
            box.show_text('Tentative de connexion...')

        if mail is not None:
            self.mail = mail
        else:
            mail = self.mail

        if pwd is not None:
            self.pwd = pwd
        else:
            pwd = self.pwd

        srv_name = 'smtp.' + mail.split('@')[1]
        self.sender_email = mail

        try:
            context = ssl.create_default_context()
            self.server = smtplib.SMTP_SSL(srv_name, 465, context=context)
            #self.server = smtplib.SMTP(srv_name)
            #self.server.ehlo()
            #self.server.starttls()
            self.server.login(mail, pwd)
            return True

        except Exception as e: # Impossible de se connecter au serveur
            global TEST_MODE
            TEST_MODE = True
            return box.warning("Connexion", ["Connexion impossible", "L'erreur suivante à été déclenchée :", str(e), "Continuer [oui] ?"]) == 'oui'

    def reconnectTest(self):
        self.counter += 1
        if self.counter > MAX_MAIL_NB:
            box.show_text('Reconnexion de sécurité en cours...')
            try:
                self.server.quit()
            except:
                self.counter = 0
                return

            time.sleep(15)
            self.connect()
            self.counter = 0

    def send(self, to, subject, text, files = [], test = True):
        """Permet d'envoyer un email automatiquement
        Arguments
        * to : destinataire
        * subject : sujet du message
        * text: corps du message
        * files : fichiers attachés au message
        * test : mode de test activé

        Retourne
        * booléen: erreur ou pas d'erreur
        """

        self.reconnectTest()
        try:
            # Création d'un mail et paramétrage des entêtes
            message = MIMEMultipart()
            message["From"] = self.sender_email
            message["To"] = to
            message["Subject"] = subject

            # Configuration du type d'affichage de text (plain vs html)
            part = MIMEText(plainToHtml(text), "html")
            message.attach(part)

            for filename in files:
                # Ouverture des fichiers joints et lecture pour encodage
                p = Path(filename)
                with open(filename, "rb") as attachment:
                    #part = MIMEBase("application", "octet-stream")
                    #part.set_payload(attachment.read())
                    part = MIMEApplication(attachment.read(), Name = str(p.name))

                # Encodage des fichiers en Base64 pour l'envoie par mail
                #encoders.encode_base64(part)

                # Ajout des descripteurs de pièce jointe
                part.add_header(
                    "Content-Disposition",
                    f"attachment; filename= {p.name}",
                )

                # Ajout de la pièce jointe
                message.attach(part)

            text = message.as_string()
            # Vérification du mode de test
            if TEST_MODE or test:
                # Simulation de l'envoie du mail, et enregistrement du fichier .eml dans le dossier temporaire
                time.sleep(0.3)
                f = open(f'temp-emails/{to}-{time.time()}.eml', 'w', encoding = 'utf-8')
                f.write(text)
                f.close()
                return True # Le mail factice a bien été envoyé

            # Envoi du mail
            self.server.sendmail(self.sender_email, to, text)
            #time.sleep(1)

            # Mail bien envoyé
            return True

        except Exception as e: # En cas d'erreur
            print()
            ask = box.warning("Envoi automatique", ['Une erreur s\'est produite en envoyant un mail.', 'Adresse de récéption prévue:', str(to), 'Erreur:', str(e), 'Voulez vous continuer ?']) == 'oui'
            if ask:
                self.connect()

            return ask


def importExcelFile(path):
    """Importation des tables de excel.
    Arguments
    * path: le fichier excel contenant les adresses mails

    Retourne
    * les listes des liste des emails
    """

    wb = xl.load_workbook(path) # Ouverture du fichier excel
    out = []
    for i, title in enumerate(['élèves', 'emplois du temps', 'appels']):
        sh = box.ask_feuille(f'Email pour les {title}', wb, path, default = i)

        ## Formatage :
        # Colonne A : groupe (si première feuille sinon on commence à la A)
        # Colonne B : nom
        # Colonne C : adresse Email
        # Colonne D : langue vivante de l'élève (facultative: None par défaut)
        # Colonne E : sous groupe fixe de l'élève !

        row = 2
        table = {}
        groupe = None
        while sh.cell(row = row, column = 2).value: # Tant qu'il y a une valeur
            if i == 0:
                # Si première table, on regarde le groupe présent ou précédent
                grp = sh.cell(row = row, column = 1).value
                if grp:
                    groupe = grp
                    table[groupe] = []

                nom = sh.cell(row = row, column = 2).value
                family = sh.cell(row = row, column = 3).value
                addr = sh.cell(row = row, column = 4).value
                lang = sh.cell(row = row, column = 5).value
                ssgrp = sh.cell(row = row, column = 6).value

                # Ajout à la table en cours
                table[groupe].append((nom, family, addr, lang, ssgrp))

            else: # Sinon, on ne regarde pas le groupe
                nom = sh.cell(row = row, column = 1).value
                addr = sh.cell(row = row, column = 2).value
                table[addr] = nom

            row += 1

        out.append(table)

    return out

def autoformat(message, variables):
    """Fonction d'application de dictionnaire sur text formaté
    A partir d'un dictionnaire et d'une str, retourne le str en
    remplacant les clé de variables par valeur

    Arguments
    * message (str): le texte de template
    * variables (dict): les valeurs à remplacer

    Retourne
    * message: le nouveau message personnalisé
    """

    for k, v in variables.items():
        message = message.replace('{' + str(k) + '}', str(v))

    return message

def ligne_colle(infos):
    """Génère un recap des colles à introduire dans le message.
    Arguments
    * infos: une liste des infos des colles

    Retourne
    * message: un text recap
    """

    txt = '<ul>'
    for salle, heure, jour, prof, groupe in infos:
        txt += f'<li>{prof} {jour.lower()} à {heure} en {salle}</li>'

    return txt + '</ul>'

def send_edt(fichiers, table_out, template, semaine, ems):
    """Envoi automatique des autres infos à qui de droit.
    Prend en charge l'envoi avec informations sur la civilité dans le
    modèle et le numéro de la semaine. Permet d'envoyer plusieurs fichier,
    à de nombreuses personne. Exemple: envoyer les listes d'appels aux profs, ...

    Arguments
    * fichiers : liste des fichiers à envoyer
    * table_out : la table des adresses mail
    * template : le fichier modèle de texte
    * semaine : le numéro de la semaine
    * ems : l'unité d'envoi des mails

    Retourne
    Rien
    """

    # Ouverture de l'envoyeur
    sujet = 'Emplois du temps semaine ' + str(semaine)

    # Récupération du modèle
    f = open(template, 'r', encoding = 'utf-8')
    content = f.read()
    f.close()

    for addr, civilite in table_out.items():
        # Envoi du mail
        ems.send(addr, sujet,
                 autoformat(content, {'civilite': civilite,
                                      'semaine': semaine,}),
                 files = fichiers,
                 test = False,
                 )

def AutoSendMail(table, files, semaine, infos, template, ems):
    """Fonction d'envoi automatique des mails pour les élèves
    Arguments
    * table : la table des adresses mail
    * files : les fichiers pdf des emplois du temps
    * semaine : le numéro de semaine actuel
    * infos : les informations sur les groupes (pour récupérer le nom de la position)
    * template : le fichier modèle de texte
    * ems: l'unité d'envoi de mails

    Retourne
    Rien
    """

    # Ouverture du serveur
    sujet = 'Emploi du temps semaine ' + str(semaine)

    # Lecture du modèle
    f = open(template, 'r', encoding = 'utf-8')
    content = f.read()
    f.close()

    l = 0
    lm = 0
    for _, eleves in table.items():
        l += len(eleves)
        for nom, _, _, _, _ in eleves:
            if len(nom) > lm:
                lm = len(nom)

    # Sécurité automatique pour ne pas envoyer par erreur tous les mails automatiquement
    ask = box.question('Confirmez la désactivation du mode de test ?', prompt = '[OUI/NON -> NON] >>> ', default = 'NON')
    if ask != 'OUI':
        TEST = True
    else:
        TEST = False

    p = box.Progress('Envoi automatique des emails', l, larg = lm)
    for groupe, eleves in table.items():
        # Récupération du fichier et de la position du groupe
        info = infos[groupe]
        for nom, family, addr, _, _ in eleves:
            fichier = files[(nom, family)]

            this_test = TEST
            if not addr: # Si il n'y a pas d'adresse mail, on fait comme si il y avait simulation d'envoi
                this_test = True

            p.step(nom, color = 'red' if not addr else None, bar = 'yellow' if this_test or TEST_MODE or not addr else None)

            # Envoi automatique
            r = ems.send(addr, sujet,
                         autoformat(content, {'name': nom,
                                              'semaine': int(semaine),
                                              'ds': semaine.DS,
                                              'colles': ligne_colle(info),
                                              'position': info[0][-1]}),

                         files = [fichier],
                         test = this_test,
                         )

            if not r: # Si une erreur s'est produite et qu'on doit s'arrêter, on coupe
                return
